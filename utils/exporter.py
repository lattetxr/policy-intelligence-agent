"""Excel 报告导出工具（政务风报告可另存为 Excel 工作簿，便于台账归档与二次编辑）。"""
import io
from pathlib import Path
from typing import List, Optional, Dict, Any

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# 政务蓝主题色
HEADER_FILL = PatternFill("solid", fgColor="165DFF")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=10, color="1D2129")
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="111824")


def _style_sheet(writer, sheet_name: str, df: pd.DataFrame, col_widths: Optional[List[int]] = None):
    ws = writer.book[sheet_name]
    header_fill = HEADER_FILL
    for col_idx, _ in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _save(workbook, path: Optional[str] = None) -> str:
    """保存到本地文件，同时返回 bytes 用于网页端下载。"""
    if path:
        Path(path).parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)
        return path
    return ""


def export_policy_elements(policies, path: str = "") -> bytes:
    """政策要素提取结果 → Excel（政策主体/申报门槛/补贴标准/量化阈值/约束条款）"""
    rows = []
    for p in policies:
        rows.append({
            "政策名称": p.title,
            "发布单位": p.issuing_body or "",
            "适用主体": "；".join(p.applicable_entities),
            "申报门槛": "；".join(p.eligibility_criteria),
            "量化阈值": "；".join(f"{k}:{v}" for k, v in p.policy_thresholds.items()),
            "补贴标准": "；".join(f"{s.subsidy_type}:{s.amount_or_ratio}" for s in p.subsidy_standards),
            "约束条款": "；".join(p.constraint_clauses),
        })
    df = pd.DataFrame(rows)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="政策要素提取")
        _style_sheet(w, "政策要素提取", df, col_widths=[26, 16, 22, 34, 30, 36, 30])
    if path:
        Path(path).parent.mkdir(parents=True, exist_ok=True)
        with open(path, "wb") as f:
            f.write(buf.getvalue())
    return buf.getvalue()


def export_match_results(ranked: List[Dict[str, Any]], path: str = "") -> bytes:
    """政企匹配结果 → Excel（政策名称/匹配度/优先级/扶持类目/适配条款/落地建议）"""
    df = pd.DataFrame(ranked)
    df = df.rename(columns={
        "rank": "排名", "policy_title": "政策名称", "match_score": "匹配度得分",
        "priority_level": "申报优先级", "policy_type": "扶持类目", "summary": "落地实操建议",
    })
    keep = [c for c in ["排名", "政策名称", "匹配度得分", "申报优先级", "扶持类目", "落地实操建议"] if c in df.columns]
    df = df[keep]
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="政企匹配结果")
        _style_sheet(w, "政企匹配结果", df, col_widths=[8, 30, 14, 14, 24, 50])
    if path:
        Path(path).parent.mkdir(parents=True, exist_ok=True)
        with open(path, "wb") as f:
            f.write(buf.getvalue())
    return buf.getvalue()


def export_application_standard(standard: Dict[str, Any], path: str = "") -> bytes:
    """申报材料标准化方案 → Excel（申报流程/材料清单/审核易错点 三个工作表）"""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        process = pd.DataFrame(standard.get("application_process", []))
        process.to_excel(w, index=False, sheet_name="申报全链路流程")
        _style_sheet(w, "申报全链路流程", process, col_widths=[24, 44, 18, 14])

        checklist = pd.DataFrame(standard.get("material_checklist", []))
        checklist.to_excel(w, index=False, sheet_name="标准化材料清单")
        _style_sheet(w, "标准化材料清单", checklist, col_widths=[22, 18, 46, 34])

        mistakes = pd.DataFrame({"审核易错点": standard.get("audit_mistakes", [])})
        mistakes.to_excel(w, index=False, sheet_name="审核易错点")
        _style_sheet(w, "审核易错点", mistakes, col_widths=[70])
    if path:
        Path(path).parent.mkdir(parents=True, exist_ok=True)
        with open(path, "wb") as f:
            f.write(buf.getvalue())
    return buf.getvalue()


def export_assessment(a, path: str = "") -> bytes:
    """政策落地评估报告 → Excel（优势/落地难点/迭代建议 工作表）"""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        overview = pd.DataFrame([{"政策名称": a.policy_title, "总体评估结论": a.overall}])
        overview.to_excel(w, index=False, sheet_name="总体评估结论")
        _style_sheet(w, "总体评估结论", overview, col_widths=[30, 70])

        strengths = pd.DataFrame([{"优势要点": s.get("title", ""), "具体说明": s.get("detail", "")} for s in a.strengths])
        strengths.to_excel(w, index=False, sheet_name="政策优势")
        _style_sheet(w, "政策优势", strengths, col_widths=[24, 70])

        difficulties = pd.DataFrame([{"落地难点": d.get("title", ""), "具体说明": d.get("detail", "")} for d in a.landing_difficulties])
        difficulties.to_excel(w, index=False, sheet_name="落地难点")
        _style_sheet(w, "落地难点", difficulties, col_widths=[24, 70])

        suggestions = pd.DataFrame({"迭代建议": a.iteration_suggestions})
        suggestions.to_excel(w, index=False, sheet_name="迭代建议")
        _style_sheet(w, "迭代建议", suggestions, col_widths=[80])
    if path:
        Path(path).parent.mkdir(parents=True, exist_ok=True)
        with open(path, "wb") as f:
            f.write(buf.getvalue())
    return buf.getvalue()
